from __future__ import annotations

import io
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

if TYPE_CHECKING:
    import sqlite3

from app.dates import format_date_de


def _try_add_pdf_logo(pdf: FPDF, width_mm: float = 26.0, y_mm: float = 10.0) -> None:
    """Platziert das WSG-Logo oben rechts auf der aktuellen PDF-Seite."""
    try:
        logo = Path(__file__).resolve().parent / "static" / "wsg-logo.png"
        if not logo.exists():
            return
        x_mm = float(pdf.w) - float(pdf.r_margin) - width_mm
        pdf.image(str(logo), x=x_mm, y=y_mm, w=width_mm)
    except Exception:
        # PDF-Erzeugung darf nicht am Logo scheitern.
        return


def _pdf_draw_report_header(
    pdf: FPDF,
    *,
    title: str,
    subtitle: str | None = None,
    logo_width_mm: float = 24.0,
) -> None:
    """Einheitlicher Kopf für alle PDF-Berichte."""
    _try_add_pdf_logo(pdf, width_mm=logo_width_mm, y_mm=10.0)
    pdf.set_font("Helvetica", "B", 15)
    pdf.cell(0, 8, _pdf_cell_text(title, 96), 0, 1)
    if subtitle:
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(0, 5.5, _pdf_cell_text(subtitle, 140), 0, 1)
    pdf.set_draw_color(184, 168, 120)
    y = float(pdf.get_y()) + 0.6
    line_start = float(pdf.l_margin)
    # Keep the decorative line clear of the top-right logo area.
    line_end = float(pdf.w) - float(pdf.r_margin) - float(logo_width_mm) - 4.0
    if line_end <= line_start + 20.0:
        line_end = float(pdf.w) - float(pdf.r_margin)
    pdf.line(line_start, y, line_end, y)
    pdf.ln(3)


def _pdf_add_page_if_needed(pdf: FPDF, needed_height_mm: float) -> bool:
    """Add a page before drawing when remaining space is insufficient."""
    if float(pdf.get_y()) + float(needed_height_mm) > float(pdf.page_break_trigger):
        pdf.add_page()
        return True
    return False


def _received_quittance_label(header: sqlite3.Row) -> str | None:
    try:
        v = header["received_confirmed"]
    except (KeyError, IndexError):
        return None
    return "Ja" if int(v) == 1 else "Nein"


def _pdf_cell_text(s: str, max_len: int) -> str:
    """PyFPDF 1.x nutzt Latin-1-Kernschriften — Umlaute ok, Rest ersetzt."""
    t = str(s).replace("\r", " ").replace("\n", " ")
    if len(t) > max_len:
        t = t[: max_len - 3] + "..."
    return t.encode("latin-1", errors="replace").decode("latin-1")


def build_xlsx_bytes(
    header: sqlite3.Row,
    aggregated_lines: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Abrechnung"
    bold = Font(bold=True)
    ws.append(["Abrechnung", "", "", ""])
    ws["A1"].font = bold
    ws.append(["Nutzer", header["user_name"]])
    ws.append(["Gruppe", header["group_name"]])
    ws.append(["Erstellt", format_date_de(header["created_at"])])
    ws.append(["Summe (EUR)", round(int(header["total_cents"]) / 100, 2)])
    if header["note"]:
        ws.append(["Notiz", header["note"]])
    rq = _received_quittance_label(header)
    if rq is not None:
        ws.append(["Zahlungseingang bestätigt", rq])
    ws.append([])
    ws.append(["Anzahl", "Bezeichnung", "Einzel (EUR)", "Summe (EUR)"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for r in aggregated_lines:
        ws.append(
            [
                int(r["quantity"]),
                r["label"],
                round(int(r["unit_cents"]) / 100, 2),
                round(int(r["total_cents"]) / 100, 2),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_bytes(
    header: sqlite3.Row,
    aggregated_lines: list[dict[str, Any]],
) -> bytes:
    """PDF ohne Pillow — PyFPDF 1.x-Paket ``fpdf`` (nicht fpdf2), keine Bild-Abhängigkeiten."""
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_margins(18, 18, 18)
    pdf.set_auto_page_break(True, 18)
    pdf.add_page()
    _pdf_draw_report_header(pdf, title="Abrechnung Shopkasse", subtitle="Nutzerabrechnung", logo_width_mm=24.0)
    pdf.set_font("Helvetica", "", 9)

    meta: list[tuple[str, str]] = [
        ("Nutzer", str(header["user_name"])),
        ("Gruppe", str(header["group_name"])),
        ("Erstellt", format_date_de(header["created_at"])),
        ("Summe EUR", f'{int(header["total_cents"]) / 100:.2f}'),
    ]
    if header["note"]:
        meta.append(("Notiz", str(header["note"])))
    rq = _received_quittance_label(header)
    if rq is not None:
        meta.append(("Zahlungseingang bestätigt", rq))

    label_w, val_w = 48, 122
    for label, val in meta:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(label_w, 6, _pdf_cell_text(label, 40), 1, 0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(val_w, 6, _pdf_cell_text(val, 120), 1, 1)

    pdf.ln(4)
    wcols = (16, 94, 30, 30)
    pdf.set_font("Helvetica", "B", 8)
    headers = ("Anz.", "Bezeichnung", "Einzel", "Summe")
    for i, (txt, w) in enumerate(zip(headers, wcols)):
        pdf.cell(w, 6, _pdf_cell_text(txt, 20), 1, 1 if i == len(wcols) - 1 else 0)
    pdf.set_font("Helvetica", "", 8)
    for r in aggregated_lines:
        cells = [
            str(int(r["quantity"])),
            str(r["label"]),
            f'{int(r["unit_cents"]) / 100:.2f}',
            f'{int(r["total_cents"]) / 100:.2f}',
        ]
        maxl = (10, 60, 12, 12)
        for i, (c, w, m) in enumerate(zip(cells, wcols, maxl)):
            pdf.cell(w, 5.5, _pdf_cell_text(c, m), 1, 1 if i == len(wcols) - 1 else 0)

    out = pdf.output(dest="S")
    return out.encode("latin-1", errors="replace")


def build_statistics_pdf_bytes(
    period_start: str | None,
    period_end: str | None,
    group_name: str | None,
    totals: dict[str, int],
    user_rows: list[dict[str, Any]],
    product_rows: list[dict[str, Any]],
    all_group_users: list[dict[str, Any]] | None = None,
) -> bytes:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_margins(14, 14, 14)
    pdf.set_auto_page_break(True, 14)
    pdf.add_page()
    _pdf_draw_report_header(pdf, title="Statistik (Zeitraum)", logo_width_mm=24.0)
    pdf.set_font("Helvetica", "", 9)
    p_from = format_date_de(period_start) if period_start else "-"
    p_to = format_date_de(period_end) if period_end else "-"
    pdf.cell(0, 6, _pdf_cell_text(f"Von: {p_from}   Bis: {p_to}", 110), 0, 1)
    if group_name:
        pdf.cell(0, 6, _pdf_cell_text(f"Nutzergruppe: {group_name}", 90), 0, 1)
    pdf.ln(1)

    meta = [
        ("Buchungen", str(int(totals["entries_count"]))),
        ("Nutzer mit Buchungen", str(int(totals["users_count"]))),
        ("Gesamtumsatz EUR", f'{int(totals["total_cents"]) / 100:.2f}'),
    ]
    for label, val in meta:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(52, 6, _pdf_cell_text(label, 36), 1, 0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(110, 6, _pdf_cell_text(val, 50), 1, 1)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, _pdf_cell_text("Nutzer-Topliste (inkl. Artikelmix)", 48), 0, 1)
    w_users = (10, 26, 28, 76, 14, 28)
    headers_u = ("#", "Gruppe", "Nutzer", "Artikel", "Anz.", "EUR")

    def _draw_users_header() -> None:
        pdf.set_font("Helvetica", "B", 8)
        for i, (txt, w) in enumerate(zip(headers_u, w_users)):
            pdf.cell(w, 6, _pdf_cell_text(txt, 20), 1, 1 if i == len(w_users) - 1 else 0)
        pdf.set_font("Helvetica", "", 8)

    _draw_users_header()
    if not user_rows:
        pdf.cell(sum(w_users), 6, _pdf_cell_text("Keine Daten im Zeitraum", 40), 1, 1)
    for idx, r in enumerate(user_rows, start=1):
        if _pdf_add_page_if_needed(pdf, 6.5):
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, _pdf_cell_text("Nutzer-Topliste (inkl. Artikelmix) - Fortsetzung", 66), 0, 1)
            _draw_users_header()
        cells = [
            str(idx),
            str(r["group_name"]),
            str(r["user_name"]),
            str(r.get("purchases_summary", "")),
            str(int(r["entries_count"])),
            f'{int(r["total_cents"]) / 100:.2f}',
        ]
        maxl = (3, 13, 14, 42, 5, 12)
        for i, (c, w, m) in enumerate(zip(cells, w_users, maxl)):
            pdf.cell(w, 5.5, _pdf_cell_text(c, m), 1, 1 if i == len(w_users) - 1 else 0)

    if group_name and all_group_users:
        _pdf_add_page_if_needed(pdf, 18.0)
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(0, 6, _pdf_cell_text("Alle Nutzer der gewaehlten Gruppe", 56), 0, 1)
        for row in all_group_users:
            article_items = cast(list[dict[str, Any]], row.get("article_items") or [])
            min_needed = 22.0 + (5.2 * max(1, len(article_items)))
            if _pdf_add_page_if_needed(pdf, min_needed):
                pdf.set_font("Helvetica", "B", 10)
                pdf.cell(0, 6, _pdf_cell_text("Alle Nutzer der gewaehlten Gruppe - Fortsetzung", 72), 0, 1)

            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, _pdf_cell_text(str(row["name"]), 42), 0, 1)
            # Wider label columns so headings are not visually clipped.
            meta_w = (42, 18, 42, 18, 42, 18)
            meta_cells = (
                "Offener Saldo",
                f'{int(row["open_balance_cents"]) / 100:.2f}',
                "Offene Buch.",
                str(int(row["open_entries_count"])),
                "Artikel ges.",
                str(int(row.get("article_count_total", 0))),
            )
            pdf.set_font("Helvetica", "B", 8)
            for i, (c, w) in enumerate(zip(meta_cells, meta_w)):
                pdf.cell(w, 6, _pdf_cell_text(c, 24), 1, 1 if i == len(meta_w) - 1 else 0)
            meta2_cells = (
                "Abgerechnet",
                f'{int(row["settled_total_cents"]) / 100:.2f}',
                "# Abrechn.",
                str(int(row["settlements_count"])),
                "",
                "",
            )
            for i, (c, w) in enumerate(zip(meta2_cells, meta_w)):
                pdf.cell(w, 6, _pdf_cell_text(c, 24), 1, 1 if i == len(meta_w) - 1 else 0)

            pdf.set_font("Helvetica", "B", 8)
            pdf.cell(132, 6, _pdf_cell_text("Artikel", 30), 1, 0)
            pdf.cell(24, 6, _pdf_cell_text("Anzahl", 12), 1, 1)
            pdf.set_font("Helvetica", "", 8)
            if not article_items:
                pdf.cell(156, 5.2, _pdf_cell_text("—", 10), 1, 1)
            else:
                for item in article_items:
                    label = str(item.get("label") or "Unbekannt")
                    qty = int(item.get("quantity") or 0)
                    if _pdf_add_page_if_needed(pdf, 6.0):
                        pdf.set_font("Helvetica", "B", 9)
                        pdf.cell(0, 6, _pdf_cell_text(str(row["name"]) + " - Fortsetzung", 54), 0, 1)
                        pdf.set_font("Helvetica", "B", 8)
                        pdf.cell(132, 6, _pdf_cell_text("Artikel", 30), 1, 0)
                        pdf.cell(24, 6, _pdf_cell_text("Anzahl", 12), 1, 1)
                        pdf.set_font("Helvetica", "", 8)
                    pdf.cell(132, 5.2, _pdf_cell_text(label, 70), 1, 0)
                    pdf.cell(24, 5.2, _pdf_cell_text(str(qty), 8), 1, 1)
            pdf.ln(1.5)

    _pdf_add_page_if_needed(pdf, 14.0)
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 6, _pdf_cell_text("Artikel (zusammengefasst)", 42), 0, 1)
    w_prod = (16, 90, 28, 28)
    headers_p = ("Anz.", "Bezeichnung", "Einzel", "Summe")

    def _draw_products_header() -> None:
        pdf.set_font("Helvetica", "B", 8)
        for i, (txt, w) in enumerate(zip(headers_p, w_prod)):
            pdf.cell(w, 6, _pdf_cell_text(txt, 20), 1, 1 if i == len(w_prod) - 1 else 0)
        pdf.set_font("Helvetica", "", 8)

    _draw_products_header()
    if not product_rows:
        pdf.cell(sum(w_prod), 6, _pdf_cell_text("Keine Daten im Zeitraum", 40), 1, 1)
    for r in product_rows:
        if _pdf_add_page_if_needed(pdf, 6.5):
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, _pdf_cell_text("Artikel (zusammengefasst) - Fortsetzung", 58), 0, 1)
            _draw_products_header()
        cells = [
            str(int(r["quantity"])),
            str(r["label"]),
            f'{int(r["unit_cents"]) / 100:.2f}',
            f'{int(r["total_cents"]) / 100:.2f}',
        ]
        for i, (c, w) in enumerate(zip(cells, w_prod)):
            pdf.cell(w, 5.5, _pdf_cell_text(c, 42), 1, 1 if i == len(w_prod) - 1 else 0)

    out = pdf.output(dest="S")
    return out.encode("latin-1", errors="replace")


def build_statistics_xlsx_bytes(
    period_start: str | None,
    period_end: str | None,
    group_name: str | None,
    totals: dict[str, int],
    user_rows: list[dict[str, Any]],
    product_rows: list[dict[str, Any]],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statistik"
    bold = Font(bold=True)

    ws.append(["Statistik (Zeitraum)"])
    ws["A1"].font = bold
    ws.append(["Von", format_date_de(period_start) if period_start else "-"])
    ws.append(["Bis", format_date_de(period_end) if period_end else "-"])
    ws.append(["Nutzergruppe", group_name or "Alle"])
    ws.append([])
    ws.append(["Kennzahl", "Wert"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    ws.append(["Buchungen", int(totals["entries_count"])])
    ws.append(["Nutzer mit Buchungen", int(totals["users_count"])])
    ws.append(["Gesamtumsatz (EUR)", round(int(totals["total_cents"]) / 100, 2)])

    ws.append([])
    ws.append(["Nutzer-Topliste"])
    ws[f"A{ws.max_row}"].font = bold
    ws.append(["Rang", "Gruppe", "Nutzer", "Artikelmix", "Buchungen", "Summe (EUR)"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for idx, r in enumerate(user_rows, start=1):
        ws.append(
            [
                idx,
                r["group_name"],
                r["user_name"],
                r.get("purchases_summary", ""),
                int(r["entries_count"]),
                round(int(r["total_cents"]) / 100, 2),
            ]
        )

    ws.append([])
    ws.append(["Artikel-Auswertung (zusammengefasst)"])
    ws[f"A{ws.max_row}"].font = bold
    ws.append(["Anzahl", "Bezeichnung", "Einzel (EUR)", "Summe (EUR)"])
    for cell in ws[ws.max_row]:
        cell.font = bold
    for r in product_rows:
        ws.append(
            [
                int(r["quantity"]),
                r["label"],
                round(int(r["unit_cents"]) / 100, 2),
                round(int(r["total_cents"]) / 100, 2),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_year_end_pdf_bytes(snapshot: dict[str, Any]) -> bytes:
    """Jahresabschluss: Kennzahlen, alle Nutzer, Artikelmix (Querformat)."""
    totals = cast(dict[str, int], snapshot["totals"])
    users = cast(list[dict[str, Any]], snapshot["users"])
    product_rows = cast(list[dict[str, Any]], snapshot["product_rows"])
    user_product_tables = cast(list[dict[str, Any]], snapshot.get("user_product_tables", []))
    created = str(snapshot["created_at_iso"])

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_margins(12, 12, 12)
    pdf.set_auto_page_break(True, 12)
    pdf.add_page()
    _pdf_draw_report_header(
        pdf,
        title="Jahresabschluss Shopkasse (Archiv)",
        subtitle=f"Erstellt (UTC): {created}",
        logo_width_mm=32.0,
    )

    meta = [
        ("Buchungszeilen gesamt", str(int(totals["ledger_entries_all"]))),
        ("Summe aller Buchungen EUR", f'{int(totals["ledger_sum_all_cents"]) / 100:.2f}'),
        ("Offene Buchungszeilen", str(int(totals["open_lines_count"]))),
        ("Offene Salden netto EUR", f'{int(totals["open_balance_net_cents"]) / 100:.2f}'),
        ("Abgeschlossene Buchungszeilen", str(int(totals["settled_lines_count"]))),
        ("Summe abgeschlossene Buchungen EUR", f'{int(totals["settled_lines_sum_cents"]) / 100:.2f}'),
        ("Anzahl Abrechnungen", str(int(totals["settlements_count"]))),
        ("Summe Abrechnungen EUR", f'{int(totals["settlements_sum_cents"]) / 100:.2f}'),
        ("Nutzer (alle)", str(int(totals["users_count"]))),
    ]
    label_w, val_w = 78, 185
    for label, val in meta:
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(label_w, 5.5, _pdf_cell_text(label, 50), 1, 0)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(val_w, 5.5, _pdf_cell_text(val, 80), 1, 1)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, _pdf_cell_text("Alle Nutzer (Stand vor Bereinigung)", 64), 0, 1)
    w_u = (38, 42, 28, 22, 22, 30, 22, 32)
    headers_u = ("Gruppe", "Nutzer", "Offener Saldo", "Offene Buch.", "# Abrechn.", "Summe Abbr.", "Buch. ges.", "Summe Buch.")

    def _draw_year_end_users_header() -> None:
        pdf.set_font("Helvetica", "B", 7)
        for i, (txt, w) in enumerate(zip(headers_u, w_u)):
            pdf.cell(w, 6, _pdf_cell_text(txt, 22), 1, 1 if i == len(w_u) - 1 else 0)
        pdf.set_font("Helvetica", "", 7)

    _draw_year_end_users_header()
    for u in users:
        if _pdf_add_page_if_needed(pdf, 6.5):
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, _pdf_cell_text("Alle Nutzer (Stand vor Bereinigung) - Fortsetzung", 76), 0, 1)
            _draw_year_end_users_header()
        cells = [
            str(u["group_name"]),
            str(u["user_name"]),
            f'{int(u["open_balance_cents"]) / 100:.2f}',
            str(int(u["open_entries_count"])),
            str(int(u["settlements_count"])),
            f'{int(u["settlements_sum_cents"]) / 100:.2f}',
            str(int(u["ledger_all_count"])),
            f'{int(u["ledger_all_sum_cents"]) / 100:.2f}',
        ]
        maxl = (18, 20, 12, 6, 6, 12, 6, 12)
        for i, (c, w, m) in enumerate(zip(cells, w_u, maxl)):
            pdf.cell(w, 5.5, _pdf_cell_text(c, m), 1, 1 if i == len(w_u) - 1 else 0)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, _pdf_cell_text("Artikel (gesamter Zeitraum, aggregiert)", 56), 0, 1)
    w_p = (18, 120, 32, 32)
    headers_p = ("Anz.", "Bezeichnung", "Einzel", "Summe")

    def _draw_year_end_products_header() -> None:
        pdf.set_font("Helvetica", "B", 8)
        for i, (txt, w) in enumerate(zip(headers_p, w_p)):
            pdf.cell(w, 6, _pdf_cell_text(txt, 20), 1, 1 if i == len(w_p) - 1 else 0)
        pdf.set_font("Helvetica", "", 8)

    _draw_year_end_products_header()
    if not product_rows:
        pdf.cell(sum(w_p), 6, _pdf_cell_text("Keine Buchungsdaten", 40), 1, 1)
    for r in product_rows:
        if _pdf_add_page_if_needed(pdf, 6.5):
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 6, _pdf_cell_text("Artikel (gesamter Zeitraum, aggregiert) - Fortsetzung", 78), 0, 1)
            _draw_year_end_products_header()
        cells = [
            str(int(r["quantity"])),
            str(r["label"]),
            f'{int(r["unit_cents"]) / 100:.2f}',
            f'{int(r["total_cents"]) / 100:.2f}',
        ]
        maxl = (8, 70, 12, 12)
        for i, (c, w, m) in enumerate(zip(cells, w_p, maxl)):
            pdf.cell(w, 5.5, _pdf_cell_text(c, m), 1, 1 if i == len(w_p) - 1 else 0)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 6, _pdf_cell_text("Artikel je Nutzer (meiste Buchungen zuerst)", 64), 0, 1)
    if not user_product_tables:
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(0, 6, _pdf_cell_text("Keine Nutzerdaten mit Buchungen", 48), 0, 1)
    for up in user_product_tables:
        _pdf_add_page_if_needed(pdf, 22.0)
        pdf.set_font("Helvetica", "B", 9)
        title = f"{up['user_name']} ({up['group_name']}) - {int(up['entries_count'])} Buchungen"
        pdf.cell(0, 6, _pdf_cell_text(title, 96), 0, 1)
        pdf.set_font("Helvetica", "", 8)
        paid_eur = f"{int(up.get('paid_cents', 0)) / 100:.2f}"
        open_eur = f"{int(up.get('open_cents', 0)) / 100:.2f}"
        total_eur = f"{int(up.get('total_cents', 0)) / 100:.2f}"
        pdf.cell(
            0,
            5.5,
            _pdf_cell_text(
                f"Gebucht/Gezahlt EUR: {paid_eur}   Offen EUR: {open_eur}   Gesamt EUR: {total_eur}",
                120,
            ),
            0,
            1,
        )
        w_up = (170, 32)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(w_up[0], 6, _pdf_cell_text("Artikel", 24), 1, 0)
        pdf.cell(w_up[1], 6, _pdf_cell_text("Anzahl", 12), 1, 1)
        pdf.set_font("Helvetica", "", 8)
        for item in cast(list[dict[str, Any]], up["items"]):
            if _pdf_add_page_if_needed(pdf, 6.0):
                pdf.set_font("Helvetica", "B", 9)
                pdf.cell(0, 6, _pdf_cell_text(title + " - Fortsetzung", 110), 0, 1)
                pdf.set_font("Helvetica", "B", 8)
                pdf.cell(w_up[0], 6, _pdf_cell_text("Artikel", 24), 1, 0)
                pdf.cell(w_up[1], 6, _pdf_cell_text("Anzahl", 12), 1, 1)
                pdf.set_font("Helvetica", "", 8)
            pdf.cell(w_up[0], 5.5, _pdf_cell_text(str(item["label"]), 78), 1, 0)
            pdf.cell(w_up[1], 5.5, _pdf_cell_text(str(int(item["quantity"])), 12), 1, 1)
        pdf.ln(1)

    out = pdf.output(dest="S")
    return out.encode("latin-1", errors="replace")


def build_year_end_xlsx_bytes(snapshot: dict[str, Any]) -> bytes:
    totals = cast(dict[str, int], snapshot["totals"])
    users = cast(list[dict[str, Any]], snapshot["users"])
    product_rows = cast(list[dict[str, Any]], snapshot["product_rows"])
    user_product_tables = cast(list[dict[str, Any]], snapshot.get("user_product_tables", []))
    created = str(snapshot["created_at_iso"])

    wb = Workbook()
    bold = Font(bold=True)

    ws0 = wb.active
    ws0.title = "Übersicht"
    ws0.append(["Jahresabschluss Shopkasse"])
    ws0["A1"].font = bold
    ws0.append(["Erstellt (UTC)", created])
    ws0.append([])
    rows_meta = [
        ("Buchungszeilen gesamt", int(totals["ledger_entries_all"])),
        ("Summe aller Buchungen (EUR)", round(int(totals["ledger_sum_all_cents"]) / 100, 2)),
        ("Offene Buchungszeilen", int(totals["open_lines_count"])),
        ("Offene Salden netto (EUR)", round(int(totals["open_balance_net_cents"]) / 100, 2)),
        ("Abgeschlossene Buchungszeilen", int(totals["settled_lines_count"])),
        ("Summe abgeschlossene Buchungen (EUR)", round(int(totals["settled_lines_sum_cents"]) / 100, 2)),
        ("Anzahl Abrechnungen", int(totals["settlements_count"])),
        ("Summe Abrechnungen (EUR)", round(int(totals["settlements_sum_cents"]) / 100, 2)),
        ("Nutzer (alle)", int(totals["users_count"])),
    ]
    ws0.append(["Kennzahl", "Wert"])
    for cell in ws0[ws0.max_row]:
        cell.font = bold
    for k, v in rows_meta:
        ws0.append([k, v])

    ws1 = wb.create_sheet("Nutzer")
    ws1.append(
        [
            "Gruppe",
            "Nutzer",
            "Offener Saldo (EUR)",
            "Offene Buchungen",
            "Anzahl Abrechnungen",
            "Summe Abrechnungen (EUR)",
            "Buchungen gesamt",
            "Summe Buchungen (EUR)",
        ]
    )
    for cell in ws1[1]:
        cell.font = bold
    for u in users:
        ws1.append(
            [
                u["group_name"],
                u["user_name"],
                round(int(u["open_balance_cents"]) / 100, 2),
                int(u["open_entries_count"]),
                int(u["settlements_count"]),
                round(int(u["settlements_sum_cents"]) / 100, 2),
                int(u["ledger_all_count"]),
                round(int(u["ledger_all_sum_cents"]) / 100, 2),
            ]
        )

    ws2 = wb.create_sheet("Artikel")
    ws2.append(["Anzahl", "Bezeichnung", "Einzel (EUR)", "Summe (EUR)"])
    for cell in ws2[1]:
        cell.font = bold
    for r in product_rows:
        ws2.append(
            [
                int(r["quantity"]),
                r["label"],
                round(int(r["unit_cents"]) / 100, 2),
                round(int(r["total_cents"]) / 100, 2),
            ]
        )

    ws3 = wb.create_sheet("Nutzer-Artikel")
    ws3.append(["Artikel je Nutzer (meiste Buchungen zuerst)"])
    ws3["A1"].font = bold
    if not user_product_tables:
        ws3.append(["Keine Nutzerdaten mit Buchungen"])
    for up in user_product_tables:
        ws3.append([])
        ws3.append([f"{up['user_name']} ({up['group_name']})", f"{int(up['entries_count'])} Buchungen"])
        ws3[f"A{ws3.max_row}"].font = bold
        ws3.append(
            [
                "Gebucht/Gezahlt (EUR)",
                round(int(up.get("paid_cents", 0)) / 100, 2),
                "Offen (EUR)",
                round(int(up.get("open_cents", 0)) / 100, 2),
                "Gesamt (EUR)",
                round(int(up.get("total_cents", 0)) / 100, 2),
            ]
        )
        ws3.append(["Artikel", "Anzahl"])
        ws3[f"A{ws3.max_row}"].font = bold
        ws3[f"B{ws3.max_row}"].font = bold
        for item in cast(list[dict[str, Any]], up["items"]):
            ws3.append([item["label"], int(item["quantity"])])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
